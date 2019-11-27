import re
import string
from openpyxl import Workbook, styles

# Initial setup
blueColor = styles.colors.Color(rgb='007B9EA8')
blueFill = styles.fills.PatternFill(patternType='solid', fgColor=blueColor)
darkColor = styles.colors.Color(rgb='0078586F')
darkFill = styles.fills.PatternFill(patternType='solid', fgColor=darkColor)
greenColor = styles.colors.Color(rgb='007FB069')
greenFill = styles.fills.PatternFill(patternType='solid', fgColor=greenColor)
whiteColor = styles.fonts.Font(color="00FFFFFF")

# Create a dictionary to store the tables
tables = {}

# Read the file tables.ddl
with open('tables.ddl', 'r') as fp:
  line = fp.readline()

  # Stores the table currently being used
  current_table = ''
  while line:
      # Check if it has already reached ALTER TABLES
      if 'ALTER TABLE ' in line:
        regex = re.findall(r'ALTER TABLE ([\w]+) ADD CONSTRAINT [\w]+ FOREIGN KEY \(([\w]+)\) REFERENCES ([\w]+) \(([\w]+)\)', line)
        if len(regex) > 0 and len(regex[0]) == 4:
          re_values = regex[0]
          if not ('fks' in tables[re_values[0]]):
            tables[re_values[0]]['fks'] = []

          tables[re_values[0]]['fks'].append(re_values[1] + "|" + re_values[2] + "." + re_values[3])

        line = fp.readline()
        continue # No need to check the rest

      # Check if it is a CREATE TABLE
      regex = re.findall(r'CREATE TABLE ([\w]+)', line)
      if len(regex) > 0:
        current_table = regex[0] # Saves the table
        tables[current_table] = {}
        tables[current_table]['attributes'] = []
        tables[current_table]['types'] = []

      # Check if it is an attribute
      regex = re.findall(r'([\w]+)\s+([\w]+)', line)
      if len(regex) > 1: # Garantees that there is a name and a type
        if regex[0] != 'PRIMARY':
          tables[current_table]['attributes'].append(regex[0][0])
          tables[current_table]['types'].append(regex[0][1])

      line = fp.readline()

# Create a new workbook
workbook = Workbook()
sheet = workbook.active

for table in tables:
  sheet.title = table

  arr_size = len(tables[table]['types'])

  # Populate tables
  for i in range(arr_size):
    cell = sheet.cell(1, i + 1, tables[table]['attributes'][i])
    sheet.column_dimensions[cell.column_letter].width = 18
    cell.fill = blueFill
    cell.font = whiteColor

    cell = sheet.cell(2, i + 1, tables[table]['types'][i])
    cell.fill = darkFill
    cell.font = whiteColor

    sheet.cell(3, i + 1, 'idk') # Placeholder

  arr_size += 3 # Padding
  # Garantees that the table has foreign keys
  fks_size = 0
  if 'fks' in tables[table]:
    fks_size = len(tables[table]['fks'])

  # Populate Foreign Keys
  for i in range(fks_size):
    fk_split = tables[table]['fks'][i].split('|')

    cell = sheet.cell(1, arr_size + i, fk_split[0])
    sheet.column_dimensions[cell.column_letter].width = 28
    cell.fill = greenFill
    cell.font = whiteColor

    cell = sheet.cell(2, arr_size + i, fk_split[1])
    cell.fill = greenFill
    cell.font = whiteColor

    # Gets the other column from the respective sheet
    original_table_split = fk_split[1].split('.')
    print(original_table_split[1])
    original_table = original_table_split[0]
    for j in range(3, 33):
      # Get the position of the cell in the other sheet
      external_pos = tables[original_table]['attributes'].index(original_table_split[1])

      external_cell = sheet.cell(j, external_pos + 1)
      sheet.cell(j, arr_size + i, '=' + original_table_split[0] + '!' + external_cell.coordinate)

  sheet = workbook.create_sheet('Sheet')

workbook.save(filename="inserts.xlsx")